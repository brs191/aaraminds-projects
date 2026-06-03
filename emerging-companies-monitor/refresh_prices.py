#!/usr/bin/env python3
"""
Live CMP refresher for ValueEducator_Portfolios_Watchlist.xlsx
==============================================================
Fetches the latest market price for every holding and writes it into the CMP
column of all sheets. The 'Live P/L %' formulas recompute automatically when you
open the file in Excel.

How a ticker is resolved (in order):
  1. NSE Symbol filled in the 'Watchlist (Combined)' sheet  -> SYMBOL.NS
  2. BSE Code filled in the 'Watchlist (Combined)' sheet    -> CODE.BO
  3. Otherwise: Yahoo Finance name search (best NSE/BSE match)
Resolved tickers are written back into the sheet so you can verify / correct
them; next run uses your value. Names that can't be resolved keep their old CMP
and are listed at the end so you can fill the ticker by hand.

Usage:
  python refresh_prices.py            # live fetch + update workbook
  python refresh_prices.py --open     # also open the file when done
  python refresh_prices.py --mock     # offline self-test (writes dummy prices)

Notes:
  - Needs 'requests' for live mode:  pip install requests
  - Close the workbook in Excel before running (Excel locks the file).
  - Prices via Yahoo Finance, personal use. Quotes may lag ~15 min; thin SME
    names may be unavailable. Not investment advice.
"""

import argparse
import sys
import time
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

WB = Path(__file__).resolve().parent / "ValueEducator_Portfolios_Watchlist.xlsx"
PORTFOLIO_SHEETS = ["Emerging Titans", "Tiny Titans"]
COMBINED = "Watchlist (Combined)"
NAME_COL, CMP_COL = 2, 8
NSE_COL, BSE_COL = 4, 5            # only on the Combined sheet
HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                         "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"}


def _get_json(url, params):
    import requests
    r = requests.get(url, params=params, headers=HEADERS, timeout=15)
    r.raise_for_status()
    return r.json()


def yahoo_search(name):
    """Resolve a company name to the best NSE/BSE Yahoo symbol, or None."""
    try:
        data = _get_json("https://query2.finance.yahoo.com/v1/finance/search",
                         {"q": name, "quotesCount": 10, "newsCount": 0})
    except Exception as e:  # noqa: BLE001
        print(f"  ! search failed for {name!r}: {e}", file=sys.stderr)
        return None
    quotes = [q for q in data.get("quotes", []) if q.get("quoteType") == "EQUITY"]
    # prefer NSE (.NS), then BSE (.BO)
    for want in (".NS", ".BO"):
        for q in quotes:
            sym = q.get("symbol", "")
            if sym.endswith(want):
                return sym
    return quotes[0].get("symbol") if quotes else None


def yahoo_price(ticker):
    """Latest regular-market price for a Yahoo ticker, or None. Tries the
    alternate exchange suffix as a fallback."""
    candidates = [ticker]
    if ticker.endswith(".NS"):
        candidates.append(ticker[:-3] + ".BO")
    elif ticker.endswith(".BO"):
        candidates.append(ticker[:-3] + ".NS")
    for t in candidates:
        try:
            j = _get_json(f"https://query1.finance.yahoo.com/v8/finance/chart/{t}",
                          {"range": "1d", "interval": "1d"})
            meta = j["chart"]["result"][0]["meta"]
            px = meta.get("regularMarketPrice")
            if px:
                return round(float(px), 2), t
        except Exception:  # noqa: BLE001
            continue
    return None, None


def get_quote(name, nse, bse, mock):
    if mock:
        return round(100 + (sum(ord(c) for c in name) % 900) + len(name) * 1.5, 1), None
    if nse:
        px, t = yahoo_price(f"{str(nse).strip()}.NS")
        if px:
            return px, t
    if bse:
        px, t = yahoo_price(f"{str(bse).strip()}.BO")
        if px:
            return px, t
    sym = yahoo_search(name)
    if sym:
        px, t = yahoo_price(sym)
        if px:
            return px, t
    return None, None


def main():
    ap = argparse.ArgumentParser(description="Refresh live CMP in the watchlist workbook")
    ap.add_argument("--mock", action="store_true", help="offline self-test, writes dummy prices")
    ap.add_argument("--open", action="store_true", help="open the workbook when done")
    args = ap.parse_args()

    if not WB.exists():
        sys.exit(f"Workbook not found: {WB}")
    try:
        wb = load_workbook(WB)
    except PermissionError:
        sys.exit("Cannot open the workbook — close it in Excel first, then re-run.")

    cw = wb[COMBINED]
    holdings = []  # (row, name, nse, bse)
    for row in range(2, cw.max_row + 1):
        name = cw.cell(row, NAME_COL).value
        if name:
            holdings.append((row, name, cw.cell(row, NSE_COL).value, cw.cell(row, BSE_COL).value))

    prices, tickers, missing = {}, {}, []
    for i, (row, name, nse, bse) in enumerate(holdings, 1):
        px, tkr = get_quote(name, nse, bse, args.mock)
        if px is None:
            missing.append(name)
            print(f"  [{i}/{len(holdings)}] {name}: not found")
            continue
        prices[name] = px
        if tkr:
            tickers[name] = tkr
        print(f"  [{i}/{len(holdings)}] {name}: {px}" + (f"  ({tkr})" if tkr else ""))
        if not args.mock:
            time.sleep(0.4)  # be gentle on the API

    # write CMP into every sheet; write resolved tickers back into Combined
    updated = 0
    for sheet in PORTFOLIO_SHEETS + [COMBINED]:
        ws = wb[sheet]
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row, NAME_COL).value
            if name in prices:
                ws.cell(row, CMP_COL).value = prices[name]
                updated += 1
    for row, name, nse, bse in holdings:
        if name in tickers:
            t = tickers[name]
            if t.endswith(".NS") and not nse:
                cw.cell(row, NSE_COL).value = t[:-3]
            elif t.endswith(".BO") and not bse:
                cw.cell(row, BSE_COL).value = t[:-3]

    stamp = datetime.now().strftime("%d %b %Y, %H:%M")
    src = "mock" if args.mock else "Yahoo Finance"
    wb["Summary"]["B3"].value = (f"{stamp}  ·  {len(prices)}/{len(holdings)} priced via {src}"
                                 + (f"  ·  {len(missing)} unresolved" if missing else ""))
    try:
        wb.save(WB)
    except PermissionError:
        sys.exit("Cannot save — close the workbook in Excel first, then re-run.")

    print(f"\nUpdated {updated} CMP cells across {len(PORTFOLIO_SHEETS)+1} sheets.")
    print(f"Priced {len(prices)}/{len(holdings)} holdings ({src}). Stamp: {stamp}")
    if missing:
        print("Unresolved (kept old CMP) — add NSE Symbol or BSE Code in the Combined sheet:")
        for m in missing:
            print("   -", m)
    print("Open the workbook in Excel; Live P/L recalculates automatically.")

    if args.open:
        import webbrowser
        webbrowser.open(WB.as_uri())


if __name__ == "__main__":
    main()
